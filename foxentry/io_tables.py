# SPDX-License-Identifier: Apache-2.0
# Copyright 2026 AVANTRO s.r.o.
"""
Reading input files and writing results.

ENCODING:
  Encoding is the most common breakage in practice. CSV exported from a Czech
  Excel is often Windows-1250 (CP1250), not UTF-8. So we detect the input:
      BOM  ->  UTF-8  ->  CP1250 (CZ/SK/PL/HU)  ->  CP1252 (W. Europe)  ->  latin-1
  A specific encoding can also be forced via INPUT_ENCODING in config.env.

  Output is written in default UTF-8 with BOM ("utf-8-sig") so Excel opens it
  with correct accents. Can be changed via OUTPUT_ENCODING.

- CSV: pure standard library (`csv`). Detects the delimiter (`,` / `;` / tab).
- XLSX: optionally via `openpyxl` (the library handles unicode itself).

Results are written CONTINUOUSLY (row by row). The app can be stopped any time
and resumed on restart - just count how many rows are already in the output.
"""

from __future__ import annotations

import re

import csv
import io
from pathlib import Path
_RE_CISELNE = re.compile(r"^[+\-]?[\d\s().\-/]+$")  # phone / number (no apostrophe needed)

try:
    import openpyxl  # type: ignore
    MA_OPENPYXL = True
except ImportError:  # pragma: no cover
    MA_OPENPYXL = False


class FileError(Exception):
    pass


# Encoding order for automatic detection. CP1250 covers CZ/SK/PL/HU,
# CP1252 western Europe. latin-1 never fails (last resort).
_AUTO_ENCODING = ["utf-8", "cp1250", "cp1252", "latin-1"]


# --------------------------------------------------------------------- READING

def _read_text(path: Path, input_encoding: str) -> tuple[str, str]:
    """Return (text, used_encoding). Detects BOM and tries candidates."""
    raw = path.read_bytes()

    # 1) Explicit BOM takes precedence
    if raw.startswith(b"\xef\xbb\xbf"):
        return raw.decode("utf-8-sig"), "utf-8 (BOM)"
    if raw.startswith(b"\xff\xfe") or raw.startswith(b"\xfe\xff"):
        return raw.decode("utf-16"), "utf-16 (BOM)"

    # 2) Forced encoding from configuration
    if input_encoding and input_encoding.lower() != "auto":
        try:
            return raw.decode(input_encoding), input_encoding
        except (LookupError, UnicodeDecodeError) as e:
            raise FileError(
                f"Could not read the file in encoding '{input_encoding}': {e}. "
                f"Zkuste INPUT_ENCODING=auto."
            )

    # 3) Automatic detection
    for enc in _AUTO_ENCODING:
        try:
            return raw.decode(enc), enc
        except UnicodeDecodeError:
            continue
    # latin-1 is in the list and never fails, we never reach here
    raise FileError("Could not detect the file encoding.")


def _detect_delimiter(sample: str) -> str:
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
        return dialect.delimiter
    except csv.Error:
        if sample.count(";") > sample.count(","):
            return ";"
        if sample.count("\t") > sample.count(","):
            return "\t"
        return ","


def load_table(path: Path, input_encoding: str = "auto") -> tuple[list[str], list[dict[str, str]], str]:
    """Load the file and return (header, rows, encoding_info)."""
    ext = path.suffix.lower()
    if ext in (".csv", ".tsv", ".txt"):
        return _load_csv(path, input_encoding)
    if ext in (".xlsx", ".xlsm"):
        h, d = _load_xlsx(path)
        from . import i18n
        return h, d, i18n.t("enc_xlsx")
    if ext == ".xls":
        raise FileError(i18n.t("err_xls_unsupported"))
    raise FileError(f"Unsupported file extension: {ext}")


def _load_csv(path: Path, input_encoding: str) -> tuple[list[str], list[dict[str, str]], str]:
    text, used = _read_text(path, input_encoding)
    # normalize line endings
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    sample = "\n".join(text.split("\n")[:5])
    delimiter = _detect_delimiter(sample)
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    rows = list(reader)
    if not rows:
        raise FileError("The file is empty.")
    header = [h.strip() for h in rows[0]]
    if not any(header):
        raise FileError("The file has no header (first row with column names).")
    data: list[dict[str, str]] = []
    for r in rows[1:]:
        if not any(cell.strip() for cell in r):
            continue
        data.append({header[i]: (r[i] if i < len(r) else "") for i in range(len(header))})
    delim_desc = {",": "sep_comma", ";": "sep_semicolon", "\t": "sep_tab"}.get(delimiter)
    from . import i18n
    sep_txt = i18n.t(delim_desc) if delim_desc else delimiter
    return header, data, f"{used}, {i18n.t('enc_sep')}: {sep_txt}"


def _load_xlsx(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    if not MA_OPENPYXL:
        raise FileError(
            "Reading .xlsx requires the openpyxl library.\n"
            "Install it with:  pip install openpyxl\n"
            "Or save the file as CSV (UTF-8)."
        )
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        first = next(rows_iter)
    except StopIteration:
        raise FileError("The sheet is empty.")
    header = [str(h).strip() if h is not None else "" for h in first]
    data: list[dict[str, str]] = []
    for r in rows_iter:
        if r is None or not any(b is not None and str(b).strip() for b in r):
            continue
        row = {}
        for i, h in enumerate(header):
            val = r[i] if i < len(r) else None
            row[h] = "" if val is None else str(val)
        data.append(row)
    wb.close()
    return header, data


# --------------------------------------------------------------------- WRITING

class StreamWriter:
    """
    Writes results continuously to CSV. Default encoding utf-8-sig (Excel + accents),
    delimiter semicolon (Czech Excel-CSV). Supports resume.
    """

    def __init__(self, path: Path, header: list[str], encoding: str = "utf-8-sig",
                 guard_csv: bool = False) -> None:
        self.path = path
        self.header = header
        self.encoding = encoding
        self.guard_csv = guard_csv
        self._f = None
        self._writer = None

    @staticmethod
    def _sanitize(value) -> object:
        """CSV/formula-injection guard - risky prefixes are prefixed with an apostrophe.
        The apostrophe is NOT added to normal data like a phone (+420...) or a negative
        number (-1500): leading + and - are dangerous only in a real formula,
        not in a purely numeric value."""
        if not isinstance(value, str) or not value:
            return value
        c = value[0]
        if c in ("=", "@", "\t", "\r"):
            return "'" + value
        if c in ("+", "-") and not _RE_CISELNE.match(value):
            return "'" + value
        return value

    def existing_rows(self) -> int:
        if not self.path.is_file():
            return 0
        with self.path.open("r", encoding=self.encoding, newline="") as f:
            count = sum(1 for _ in f)
        return max(0, count - 1)

    def load_results(self, result_columns, not_filled, limit=None):
        """Read the result breakdown from the already-written output (for the report after resume).

        Returns (Counter by_result, api_calls, errors). Counted per cell
        (service x row); empty/unfilled ones are not counted (no call happened).
        """
        from collections import Counter
        from . import i18n
        out: Counter = Counter()
        calls = 0
        errors = 0
        if not self.path.is_file():
            return out, calls, errors
        err = i18n.t("res_error")
        with self.path.open("r", encoding=self.encoding, newline="") as f:
            rdr = csv.reader(f, delimiter=";")
            try:
                hdr = next(rdr)
            except StopIteration:
                return out, calls, errors
            idx = {c.strip(): i for i, c in enumerate(hdr)}
            cols = [idx[c] for c in result_columns if c in idx]
            rows_n = 0
            for row in rdr:
                if limit is not None and rows_n >= limit:
                    break
                if not any(c.strip() for c in row):
                    continue
                rows_n += 1
                for ci in cols:
                    v = row[ci].strip() if ci < len(row) else ""
                    if not v or v == not_filled:
                        continue
                    out[v] += 1
                    if v == err:
                        errors += 1
                    else:
                        calls += 1
        return out, calls, errors

    def header_matches(self) -> bool:
        if not self.path.is_file():
            return True
        with self.path.open("r", encoding=self.encoding, newline="") as f:
            first = f.readline().rstrip("\r\n")
        # the CSV writer may quote - compare parsed
        try:
            existing = next(csv.reader([first], delimiter=";"))
        except StopIteration:
            existing = []
        return [c.strip() for c in existing] == [c.strip() for c in self.header]

    def open_writer(self, resume: bool) -> None:
        mode = "a" if (resume and self.path.is_file()) else "w"
        is_new = mode == "w"
        self._f = self.path.open(mode, encoding=self.encoding, newline="")
        self._writer = csv.writer(self._f, delimiter=";", quoting=csv.QUOTE_MINIMAL)
        if is_new:
            self._writer.writerow(self.header)
            self._f.flush()

    def write_row(self, row: dict[str, object]) -> None:
        assert self._writer is not None and self._f is not None
        values = [row.get(h, "") for h in self.header]
        if self.guard_csv:
            values = [self._sanitize(v) for v in values]
        self._writer.writerow(values)
        self._f.flush()  # immediately to disk -> crash-safe

    def close_writer(self) -> None:
        if self._f is not None:
            self._f.close()
            self._f = None
            self._writer = None


def csv_to_xlsx(csv_path: Path, xlsx_path: Path, encoding: str = "utf-8-sig") -> bool:
    """After completion, convert the output CSV to .xlsx (if openpyxl is available)."""
    if not MA_OPENPYXL:
        return False
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Vysledek"
    with csv_path.open("r", encoding=encoding, newline="") as f:
        for row in csv.reader(f, delimiter=";"):
            ws.append(row)
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
    ws.freeze_panes = "A2"
    wb.save(xlsx_path)
    wb.close()
    return True


def no_header(header: list[str], rows: list[dict[str, str]]):
    """The first physical row is NOT column names -> generate col1.. and shift data."""
    n = len(header)
    new_cols = [f"col{i + 1}" for i in range(n)]
    first = {new_cols[i]: header[i] for i in range(n)}
    nxt = [{new_cols[i]: r.get(header[i], "") for i in range(n)} for r in rows]
    return new_cols, [first] + nxt
